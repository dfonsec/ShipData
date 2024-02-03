
import openpyxl
from openpyxl import load_workbook

wb = openpyxl.load_workbook("/Users/danielfonseca/repos/ShipData/data/ship_maintence.xlsx")

def deleteDuplicateRows(workbook):
    """
    Deletes duplicate rows based on unique Service Order IDs in the 'Ship_Maintenance_Cleaned' sheet.
    Creates a new sheet named 'No Duplicates' and appends rows with distinct Service Order IDs.

    Args:
    - workbook: An openpyxl Workbook object containing the data.

    Returns:
    - The modified workbook with the 'No Duplicates' sheet.
    - Saves the result to a new Excel file ('resultData.xlsx').
    """
    
    ws=workbook['Ship_Maintenance_Cleaned']
    rows = ws.iter_rows(min_row=1)
    seenIDS = set()
    resultSheet = workbook.create_sheet('No Duplicates')
    
    for row in rows:
        if row[2].value not in seenIDS:
            seenIDS.add(row[2].value)
            row_with_values = [cell.value for cell in row]
            resultSheet.append(row_with_values)
    
    return workbook.save("/Users/danielfonseca/repos/ShipData/data/resultData.xlsx")

def countUniqueRows(worksheet):
    """
    Counts the number of rows with unique Service Order IDs in the provided worksheet.

    Args:
    - worksheet: An openpyxl Worksheet object containing the data.

    Returns:
    - The count of rows with unique Service Order IDs.
    """
    rows = worksheet.iter_rows(min_row=2)
    seenIDS = set()

    for row in rows:
        if row[2].value not in seenIDS:
            seenIDS.add(row[2].value)

    return len(seenIDS)
            


    


            
            
    





    
    
    
    



